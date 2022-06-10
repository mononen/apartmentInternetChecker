import openpyxl
import attRequest


def clean_address(address):
  addrLn = address.split(",")[0]
  zipcode = address[-5:]
  return addrLn, zipcode

def writeResults(res, addr, sheet, wb):
  if res["status"] == "No Offers" or "ERROR" in res["status"]:
    sheet.cell(row=addr, column=1).value = res["status"]
    print(res["status"])
  else:
    sheet.cell(row=addr, column=1).value = res["status"]
    sheet.cell(row=addr, column=4).value = res['provider']
    sheet.cell(row=addr, column=6).value = res['maxDownload']
    sheet.cell(row=addr, column=8).value = res['packageName']
  wb.save("data.xlsx")

loc = "./data.xlsx"

wb = openpyxl.load_workbook(loc)
sheet = wb.active
m_row = sheet.max_row

for addr in range(1, m_row + 1):
  cell = sheet.cell(row=addr, column=3).value
  statusCell = sheet.cell(row=addr, column=1).value
  if cell[-3].isnumeric() == True:
    if statusCell != "success" and statusCell != "No Offers":
      addrLn, zipcode = clean_address(cell)
      print(addrLn, zipcode)
      res = attRequest.checkAddress(addrLn, zipcode)
      writeResults(res, addr, sheet, wb)
    else:
      print("Already checked")
  else:
    print("No zipcode")


